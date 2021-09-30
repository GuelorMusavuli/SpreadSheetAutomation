# Learn how to work with Spreadsheets by
# reading Spreadsheet file and
# automate stuff in case they are many files (do processing on data)

import openpyxl

# Read the entire spreadsheet's content
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]  # access a particular sheet in the file

products_per_supplier = {}  # {"company" : product}
total_value_per_supplier = {}  # {"company" : total value}
products_under_10_inventory = {}  # {"product number" : inventory value}

# Iterate through each and every item in the sheet.
for product_row in range(2, product_list.max_row + 1):

    # Get the values within the sheet(row, column).
    product_number = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)  # to write to the spreadsheet

    # 1. Calculation number of products per supplier :
    # if an existing supplier in the dictionary, increment the product value associated with it.
    # else, add the product value with the newly created supplier.
    if supplier_name in products_per_supplier:
        current_products_number = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_products_number + 1
    else:
        products_per_supplier[supplier_name] = 1

    # 2. Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # 3. Calculation products with inventory less than 10
    if inventory < 10:
        products_under_10_inventory[int(product_number)] = int(inventory)

    # 4. Calculation to add value for total inventory price for each product into the file
    inventory_price.value = inventory * price


# List each company with respective product count
print(products_per_supplier)

# List products with inventory less than 10
print(products_under_10_inventory)

# List each company with respective total inventory value
print(total_value_per_supplier)

# Save changes and create a new file from the original file
inv_file.save("inventory_with_total_value.xlsx")
